import openpyxl
import os

# 获取Excel文件路径
file_path = os.path.join(os.getcwd(), '车辆规划车辆收集生成甘特图模板.xlsm')
print(f"分析文件: {file_path}")

# 打开Excel文件
workbook = openpyxl.load_workbook(file_path, data_only=True)

# 获取工作表列表
print("\n工作表列表:")
for sheet_name in workbook.sheetnames:
    print(f"- {sheet_name}")

# 分析每个工作表的内容
for sheet_name in workbook.sheetnames:
    print(f"\n\n=== 工作表: {sheet_name} ===")
    sheet = workbook[sheet_name]
    
    # 获取最大行和列
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"行数: {max_row}, 列数: {max_col}")
    
    # 打印前10行数据（如果有）
    print("\n前10行数据:")
    for row in range(1, min(max_row + 1, 11)):
        row_data = []
        for col in range(1, min(max_col + 1, 11)):  # 只显示前10列
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(str(cell_value) if cell_value is not None else "")
        print(f"第{row}行: {row_data}")

# 检查是否有宏
print("\n\n=== 宏信息 ===")
# 尝试读取VBA项目信息
try:
    # 检查是否存在VBA项目
    if hasattr(workbook, 'vba_archive'):
        print("文件包含VBA宏")
    else:
        print("文件不包含VBA宏")
except Exception as e:
    print(f"检查宏时出错: {e}")

workbook.close()
